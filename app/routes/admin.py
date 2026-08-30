import requests
import os
import json
from flask import Blueprint, render_template, redirect, url_for, request, flash, session, jsonify, Response, send_file
from app.utils import BASE_URL, get_headers, role_required
from app.api_helpers import extract_list, with_list_key
import io
from datetime import datetime
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

admin_bp = Blueprint('admin', __name__)

# --- HELPDESK ---
@admin_bp.route('/helpdesk')
@role_required(['admin', 'employee', 'hr', 'manager'])
def helpdesk():
    return render_template('helpdesk.html', BASE_URL=BASE_URL)

# --- SOFTWARE ---
@admin_bp.route('/software')
@role_required(['admin', 'employee', 'hr', 'manager'])
def software():
    return render_template('software.html', BASE_URL=BASE_URL)

@admin_bp.route('/api/software', methods=['GET', 'POST'])
@role_required(['admin'])
def api_software():
    if request.method == 'POST':
        res = requests.post(f"{BASE_URL}/software", json=request.get_json(), headers=get_headers())
    else:
        res = requests.get(f"{BASE_URL}/software", headers=get_headers())
    return jsonify(res.json()), res.status_code

@admin_bp.route('/api/software/<int:software_id>', methods=['GET', 'PUT', 'DELETE'])
@role_required(['admin'])
def api_software_detail(software_id):
    if request.method == 'PUT':
        res = requests.put(f"{BASE_URL}/software/{software_id}", json=request.get_json(), headers=get_headers())
    elif request.method == 'DELETE':
        res = requests.delete(f"{BASE_URL}/software/{software_id}", headers=get_headers())
    else:
        res = requests.get(f"{BASE_URL}/software/{software_id}", headers=get_headers())
    return jsonify(res.json()), res.status_code

@admin_bp.route('/api/helpdesk/', methods=['GET', 'POST'])
def api_helpdesk_list():
    if request.method == 'POST':
        res = requests.post(f"{BASE_URL}/helpdesk/", json=request.get_json(), headers=get_headers())
    else:
        res = requests.get(f"{BASE_URL}/helpdesk/", params=request.args.to_dict(), headers=get_headers())
        try:
            body = with_list_key(res.json(), 'tickets', 'data')
            return jsonify(body), res.status_code
        except Exception:
            pass
    return jsonify(res.json()), res.status_code

@admin_bp.route('/api/helpdesk/<int:ticket_id>', methods=['GET'])
def api_helpdesk_detail(ticket_id):
    res = requests.get(f"{BASE_URL}/helpdesk/{ticket_id}", headers=get_headers())
    return jsonify(res.json()), res.status_code

@admin_bp.route('/api/helpdesk/<int:ticket_id>/status', methods=['PATCH'])
@role_required(['admin', 'hr'])
def api_helpdesk_status(ticket_id):
    res = requests.patch(f"{BASE_URL}/helpdesk/{ticket_id}/status", json=request.get_json(), headers=get_headers())
    return jsonify(res.json()), res.status_code

@admin_bp.route('/api/helpdesk/<int:ticket_id>/assign', methods=['PATCH'])
@role_required(['admin', 'hr'])
def api_helpdesk_assign(ticket_id):
    res = requests.patch(f"{BASE_URL}/helpdesk/{ticket_id}/assign", json=request.get_json(), headers=get_headers())
    return jsonify(res.json()), res.status_code

@admin_bp.route('/reimbursement')
@admin_bp.route('/reimbursements')
@role_required(['admin', 'employee', 'hr', 'manager'])
def reimbursement():
    return render_template('reimbursement.html', BASE_URL=BASE_URL)

@admin_bp.route('/api/reimbursements', methods=['GET', 'POST'])
@role_required(['admin', 'employee', 'hr', 'manager'])
def api_reimbursements():
    if request.method == 'POST':
        # Handle multipart form data for file upload
        if request.content_type and 'multipart' in request.content_type:
            form_data = request.form.to_dict()
            files = {}
            if 'receipt' in request.files:
                receipt_file = request.files['receipt']
                if receipt_file.filename:
                    files['receipt'] = (receipt_file.filename, receipt_file.read(), receipt_file.content_type)
            
            res = requests.post(f"{BASE_URL}/reimbursements/", data=form_data, files=files, headers=get_headers(exclude_content_type=True))
        else:
            res = requests.post(f"{BASE_URL}/reimbursements/", json=request.get_json(), headers=get_headers())
    else:
        res = requests.get(f"{BASE_URL}/reimbursements/", params=request.args.to_dict(), headers=get_headers())
        try:
            body = with_list_key(res.json(), 'reimbursements', 'data')
            return jsonify(body), res.status_code
        except Exception:
            pass
    
    return jsonify(res.json()), res.status_code

@admin_bp.route('/api/reimbursements/<int:record_id>', methods=['GET', 'DELETE'])
@role_required(['admin', 'employee', 'hr', 'manager', 'accounts', 'superadmin'])
def api_reimbursement_detail(record_id):
    if request.method == 'DELETE':
        res = requests.delete(f"{BASE_URL}/reimbursements/{record_id}", headers=get_headers())
    else:
        res = requests.get(f"{BASE_URL}/reimbursements/{record_id}", headers=get_headers())
        if res.status_code == 200:
            try:
                data = res.json()
                if not data.get('history'):
                    h_res = requests.get(f"{BASE_URL}/reimbursements/{record_id}/history", headers=get_headers())
                    if h_res.status_code == 200:
                        data['history'] = h_res.json().get('history', [])
                return jsonify(data)
            except Exception as e:
                logging.error(f"Error parsing reimbursement detail: {e}")
    try:
        return jsonify(res.json()), res.status_code
    except Exception:
        return jsonify({"success": False, "error": "Failed to fetch reimbursement details"}), res.status_code

@admin_bp.route('/api/reimbursements/<int:record_id>/approve', methods=['PATCH'])
@role_required(['admin', 'hr', 'manager'])
def api_approve_reimbursement(record_id):
    res = requests.patch(f"{BASE_URL}/reimbursements/{record_id}/approve", json=request.get_json() or {}, headers=get_headers())
    return jsonify(res.json()), res.status_code

@admin_bp.route('/api/reimbursements/<int:record_id>/reject', methods=['PATCH'])
@role_required(['admin', 'hr', 'manager'])
def api_reject_reimbursement(record_id):
    res = requests.patch(f"{BASE_URL}/reimbursements/{record_id}/reject", json=request.get_json() or {}, headers=get_headers())
    return jsonify(res.json()), res.status_code

@admin_bp.route('/api/reimbursements/<int:record_id>/pay', methods=['PATCH'])
@role_required(['admin'])
def api_pay_reimbursement(record_id):
    res = requests.patch(f"{BASE_URL}/reimbursements/{record_id}/pay", json=request.get_json() or {}, headers=get_headers())
    return jsonify(res.json()), res.status_code

@admin_bp.route('/api/reimbursements/<int:record_id>/receipt', methods=['GET'])
@role_required(['admin', 'employee', 'hr', 'manager'])
def api_reimbursement_receipt(record_id):
    res = requests.get(f"{BASE_URL}/reimbursements/{record_id}/receipt", headers=get_headers(), stream=True)
    from flask import Response
    return Response(res.content, status=res.status_code, headers=dict(res.headers))

# --- ASSETS ---
@admin_bp.route('/assets')
@role_required(['admin', 'hr'])
def assets():
    return render_template('assets.html', BASE_URL=BASE_URL)

# --- POLICIES ---
@admin_bp.route('/policies')
@role_required(['admin', 'employee', 'hr', 'manager'])
def policies():
    res = requests.get(f"{BASE_URL}/reports/policies", headers=get_headers())
    data = res.json() if res.status_code == 200 else {}
    policies_list = extract_list(data, 'policies', 'data')
    categories = sorted(list(set(p.get('category', 'General') for p in policies_list)))
    return render_template("policies.html", policies=policies_list, categories=categories, BASE_URL=BASE_URL)

@admin_bp.route('/api/policies/<int:policy_id>', methods=['PUT'])
@role_required(['admin', 'hr'])
def api_update_policy(policy_id):
    try:
        data = request.get_json()
        res = requests.put(
            f"{BASE_URL}/reports/policies/{policy_id}",
            json=data,
            headers=get_headers()
        )
        if res.status_code == 200:
            return jsonify(res.json()), 200
        else:
            try:
                return jsonify(res.json()), res.status_code
            except:
                return jsonify({"success": False, "error": res.text}), res.status_code
    except Exception as e:
        print(f"ERROR in policy update API: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@admin_bp.route('/bank-details/<int:detail_id>/<action>', methods=['PATCH'])
@role_required(['admin', 'hr'])
def verify_bank_admin(detail_id, action):
    if action not in ['approve', 'reject']:
        return jsonify({"success": False, "error": "Invalid action"}), 400
    try:
        res = requests.patch(f"{BASE_URL}/bank-details/{detail_id}/{action}", 
                           json=request.get_json() or {}, headers=get_headers(), timeout=5)
        return jsonify(res.json()), res.status_code
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@admin_bp.route('/api/assets', methods=['GET', 'POST'])
@role_required(['admin', 'hr'])
def api_assets():
    if request.method == 'GET':
        try:
            res = requests.get(f"{BASE_URL}/devices", headers=get_headers(), timeout=10)
            if res.status_code == 200:
                data = res.json()
                assets = data if isinstance(data, list) else extract_list(data, 'assets', 'devices', 'data')
                return jsonify({"success": True, "assets": assets})
            return jsonify({"success": False, "error": res.text}), res.status_code
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500
    
    else: # POST
        try:
            # Always extract text fields into a clean dict regardless of content type
            # The backend /devices API expects JSON — never send raw form-data to it
            if request.content_type and 'multipart' in request.content_type:
                payload = request.form.to_dict()
            else:
                payload = request.get_json(force=True) or {}

            # Build clean JSON payload — skip None/empty values
            create_data = {k: v for k, v in {
                "device_name": payload.get("device_name"),
                "device_type": payload.get("device_type"),
                "serial_number": payload.get("serial_number"),
                "asset_id": payload.get("asset_id"),
                "brand": payload.get("brand") or payload.get("device_name"),
                "model": payload.get("model") or payload.get("device_type"),
                "status": "Available",
                "processor": payload.get("processor"),
                "ram": payload.get("ram"),
                "storage": payload.get("storage"),
                "purchase_date": payload.get("purchase_date"),
                "warranty_expiry": payload.get("warranty_expiry"),
                "condition_notes": payload.get("notes"),
                # Ownership / rental details
                "ownership_type": payload.get("ownership_type", "Purchased"),
                "vendor_name": payload.get("vendor_name"),
                "vendor_contact": payload.get("vendor_contact"),
                "rental_start_date": payload.get("rental_start_date"),
                "rental_end_date": payload.get("rental_end_date"),
                "rental_cost": payload.get("rental_cost"),
                "rental_cost_frequency": payload.get("rental_cost_frequency"),
            }.items() if v}

            # Step 1: Create Device — always send as JSON
            res = requests.post(f"{BASE_URL}/devices", json=create_data, headers=get_headers(), timeout=10)

            if res.status_code not in [200, 201]:
                try:
                    err_detail = res.json()
                except Exception:
                    err_detail = res.text
                return jsonify({"success": False, "error": f"Creation failed: {err_detail}"}), res.status_code

            data = res.json()
            device_id = data.get("device_id") or data.get("id")

            # Step 2: Upload images if present (separate request per file)
            if device_id and request.files:
                img_headers = get_headers(exclude_content_type=True)
                for key in request.files:
                    for file in request.files.getlist(key):
                        if not file or not file.filename:
                            continue
                        file.seek(0)
                        f_data = {'image': (file.filename, file.read(), file.content_type)}
                        requests.post(
                            f"{BASE_URL}/devices/{device_id}/upload-image",
                            files=f_data,
                            headers=img_headers,
                            timeout=15
                        )
                        break  # Backend supports one primary image

            return jsonify({
                "success": True,
                "asset": {
                    "id": device_id,
                    "device_name": create_data.get("device_name"),
                    "serial_number": create_data.get("serial_number")
                }
            })

        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500

@admin_bp.route('/api/assets/export', methods=['GET'])
@role_required(['admin', 'hr'])
def export_assets():
    try:
        # Fetch assets
        res = requests.get(f"{BASE_URL}/devices", headers=get_headers(), timeout=15)
        if res.status_code != 200:
            return jsonify({"success": False, "error": f"Failed to fetch assets: {res.text}"}), res.status_code
        
        data = res.json()
        assets = data if isinstance(data, list) else extract_list(data, 'assets', 'devices', 'data')
        
        if not assets:
            return jsonify({"success": False, "error": "No assets found to export"}), 404

        # Query Parameters for Filtering
        export_type = request.args.get('export_type', 'all')
        category = request.args.get('category')
        status = request.args.get('status')
        ownership = request.args.get('ownership')
        vendor = request.args.get('vendor')
        sort_by = request.args.get('sort_by', 'device_name')
        sort_dir = request.args.get('sort_dir', 'asc')

        filtered_assets = []
        for a in assets:
            # Determine fields
            a_status = (a.get('status') or '').lower()
            a_emp_id = a.get('employee_id') or a.get('assigned_to') or a.get('user_id') or a.get('emp_id') or a.get('allotted_to') or a.get('allotted_id')
            is_assigned = (a_status in ['assigned', 'allotted']) or bool(a_emp_id)
            a_ownership = a.get('ownership_type', 'Purchased')

            # Filter by Export Type
            if export_type == 'available' and is_assigned: continue
            if export_type == 'assigned' and not is_assigned: continue
            if export_type == 'purchased' and a_ownership != 'Purchased': continue
            if export_type == 'rented' and a_ownership != 'Rented': continue
            if export_type == 'repair' and a_status != 'under repair': continue
            if export_type == 'retired' and a_status != 'retired': continue
            
            # Additional Optional Filters
            if category and (a.get('device_type') or '') != category: continue
            if status and a_status != status.lower(): continue
            if ownership and a_ownership != ownership: continue
            if vendor and a.get('vendor_name') != vendor: continue
            
            filtered_assets.append(a)

        if not filtered_assets:
            return jsonify({"success": False, "error": "No matching assets found for the selected filters"}), 404

        # Sorting
        reverse_sort = (sort_dir == 'desc')
        def sort_key(a):
            if sort_by == 'id': return a.get('id') or a.get('device_id') or 0
            if sort_by == 'tag': return a.get('asset_id') or ''
            if sort_by == 'purchase_date': return a.get('purchase_date') or ''
            if sort_by == 'vendor': return a.get('vendor_name') or ''
            if sort_by == 'status': return a.get('status') or ''
            if sort_by == 'category': return a.get('device_type') or ''
            return a.get('device_name') or a.get('brand') or ''
        
        filtered_assets.sort(key=sort_key, reverse=reverse_sort)

        # Audit Logging
        user_id = session.get('user_id')
        user_name = session.get('name')
        role = session.get('role')
        ip_addr = request.remote_addr
        logging.info(f"AUDIT: Asset Export | User: {user_name} ({user_id}, {role}) | IP: {ip_addr} | Filters: {request.args.to_dict()} | Total Exported: {len(filtered_assets)}")

        # Create Excel File
        wb = Workbook()
        ws = wb.active
        ws.title = "Asset Inventory"

        # Headers
        headers = [
            "Asset ID", "Asset Tag", "Asset Name", "Category", "Ownership Type", 
            "Status", "Model", "Serial Number", "Vendor Name", "Purchase Date", 
            "Purchase Price", "Warranty Expiry", "Current Assigned Team Member", 
            "Created Date", "Last Updated"
        ]
        
        # Write Report Metadata
        ws.append(["Company Name:", "Altzor", "", "", "", "Report Title:", "Asset Inventory Report"])
        ws.append(["Generated By:", f"{user_name} ({role})", "", "", "", "Generated Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws.append(["Total Records:", str(len(filtered_assets)), "", "", "", "Filters Applied:", str(dict(request.args))])
        ws.append([]) # Empty row

        # Write Headers
        ws.append(headers)
        header_row = 5
        
        # Style Headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F46E5")
        for col_num, cell in enumerate(ws[header_row], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = f"A{header_row + 1}"
        ws.auto_filter.ref = f"A{header_row}:O{header_row}"

        # Write Data
        for a in filtered_assets:
            a_emp_id = a.get('employee_id') or a.get('assigned_to') or a.get('user_id') or a.get('emp_id') or a.get('allotted_to') or a.get('allotted_id')
            a_status = a.get('status') or 'Available'
            if a_emp_id and a_status.lower() not in ['assigned', 'allotted']:
                a_status = 'Assigned'
                
            row = [
                a.get('id') or a.get('device_id') or '',
                a.get('asset_id') or '',
                a.get('device_name') or a.get('brand') or '',
                a.get('device_type') or '',
                a.get('ownership_type', 'Purchased'),
                a_status.capitalize(),
                a.get('model') or a.get('device_type') or '',
                a.get('serial_number') or '',
                a.get('vendor_name') or '',
                a.get('purchase_date') or '',
                a.get('rental_cost') or '', # using rental_cost as purchase_price mapping for simplicity if available
                a.get('warranty_expiry') or a.get('warranty_end') or '',
                a_emp_id or '',
                a.get('created_at') or '',
                a.get('updated_at') or ''
            ]
            ws.append(row)

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 50)

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Asset_Inventory_{datetime.now().strftime('%Y_%m_%d_%H_%M')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logging.error(f"Asset Export Error: {str(e)}")
        return jsonify({"success": False, "error": "An error occurred while generating the Excel file."}), 500

@admin_bp.route('/api/assets/<id>', methods=['GET', 'DELETE', 'PUT'])
@role_required(['admin', 'hr'])
def api_asset_detail(id):
    if request.method == 'GET':
        try:
            res = requests.get(f"{BASE_URL}/devices/{id}", headers=get_headers(), timeout=10)
            if res.status_code == 200:
                return jsonify({"success": True, "asset": res.json()})
            return jsonify({"success": False, "error": res.text}), res.status_code
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500

    elif request.method == 'PUT':
        try:
            payload = request.get_json(force=True) or {}
            res = requests.put(f"{BASE_URL}/devices/{id}", json=payload, headers=get_headers(), timeout=10)
            if res.status_code in [200, 201, 204]:
                return jsonify({"success": True})
            return jsonify({"success": False, "error": res.text}), res.status_code
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500

    else: # DELETE
        try:
            res = requests.delete(f"{BASE_URL}/devices/{id}", headers=get_headers(), timeout=10)
            if res.status_code in [200, 204]:
                return jsonify({"success": True})
            return jsonify({"success": False, "error": res.text}), res.status_code
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500

@admin_bp.route('/api/assets/<id>/history')
@role_required(['admin', 'hr'])
def api_asset_history(id):
    try:
        res = requests.get(f"{BASE_URL}/devices/{id}/history", headers=get_headers(), timeout=10)
        if res.status_code == 200:
            return jsonify({"success": True, "history": res.json()})
        return jsonify({"success": False, "error": res.text}), res.status_code
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@admin_bp.route('/api/assets/<id>/assign', methods=['POST'])
@role_required(['admin', 'hr'])
def api_assign_asset(id):
    try:
        data = request.get_json(force=True) or {}
        # Only pass non-empty values to the backend
        assign_payload = {k: v for k, v in data.items() if v not in [None, '', []]}
        res = requests.post(f"{BASE_URL}/devices/{id}/assign", json=assign_payload, headers=get_headers(), timeout=10)
        if res.status_code in [200, 201]:
            return jsonify({"success": True})
        try:
            err_detail = res.json()
        except Exception:
            err_detail = res.text
        return jsonify({"success": False, "error": err_detail}), res.status_code
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@admin_bp.route('/api/assets/<id>/return', methods=['POST'])
@role_required(['admin', 'hr'])
def api_return_asset(id):
    try:
        res = requests.post(f"{BASE_URL}/devices/{id}/return", headers=get_headers(), timeout=10)
        if res.status_code in [200, 201]:
            return jsonify({"success": True, "message": "Asset returned successfully", "asset_status": "AVAILABLE"})
        try:
            err_detail = res.json()
        except Exception:
            err_detail = res.text
        return jsonify({"success": False, "error": err_detail}), res.status_code
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@admin_bp.route('/api/assets/<id>/acceptance-status')
@role_required(['admin', 'hr'])
def api_asset_acceptance(id):
    try:
        res = requests.get(f"{BASE_URL}/devices/{id}/acceptance-status", headers=get_headers(), timeout=10)
        if res.status_code == 200:
            return jsonify({"success": True, "status": res.json()})
        return jsonify({"success": False, "error": res.text}), res.status_code
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- AGREEMENT PAGE ---
@admin_bp.route('/assets/agreement/<id>')
@role_required(['admin', 'hr', 'manager', 'employee', 'team_member'])
def asset_agreement_page(id):
    """Render the device usage agreement signing page."""
    from app.ui_constants import UI_LABELS
    return render_template('agreement.html', asset_id=id, labels=UI_LABELS, BASE_URL=BASE_URL)

@admin_bp.route('/api/assets/<id>/agreement')
@role_required(['admin', 'hr', 'manager', 'employee', 'team_member'])
def api_asset_agreement(id):
    """Fetch agreement/assignment details for a device."""
    try:
        res = requests.get(f"{BASE_URL}/devices/{id}", headers=get_headers(), timeout=10)
        if res.status_code != 200:
            return jsonify({"success": False, "error": res.text}), res.status_code
        data = res.json()
        # Flatten: backend may nest inside .device
        device = data.get('device') or data
        agreement = {
            "assignment_id": device.get('assignment_id') or device.get('id'),
            "assigned_date": device.get('assigned_date') or device.get('created_at', ''),
            "employee_name": device.get('employee_name') or device.get('assigned_to', ''),
            "employee_id": device.get('employee_id', ''),
            "device": {
                "brand": device.get('brand') or device.get('device_name', ''),
                "model": device.get('model') or device.get('device_type', ''),
                "serial_number": device.get('serial_number', ''),
            }
        }
        return jsonify({"success": True, "agreement": agreement})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@admin_bp.route('/api/assets/<id>/accept', methods=['POST'])
@role_required(['admin', 'hr', 'manager', 'employee', 'team_member'])
def api_accept_agreement(id):
    """Submit signed agreement."""
    try:
        files = {}
        if 'signature' in request.files:
            sig = request.files['signature']
            if sig.filename:
                files['signature'] = (sig.filename, sig.read(), sig.content_type)

        form_data = request.form.to_dict()
        
        if files:
            res = requests.post(
                f"{BASE_URL}/devices/{id}/accept",
                data=form_data,
                files=files,
                headers=get_headers(exclude_content_type=True),
                timeout=15
            )
        else:
            res = requests.post(
                f"{BASE_URL}/devices/{id}/accept",
                data=form_data,
                headers=get_headers(exclude_content_type=True),
                timeout=15
            )
            
        try:
            body = res.json()
        except Exception:
            body = {"success": res.status_code in [200, 201]}
        return jsonify(body), res.status_code
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# --- ANNOUNCEMENTS ---


@admin_bp.route('/api/announcements', methods=['GET', 'POST'])
@role_required(['admin', 'employee', 'hr', 'manager'])
def api_announcements():
    if request.method == 'POST':
        if str(session.get('role', '')).lower().strip() != 'hr':
            return jsonify({"success": False, "error": "Access denied"}), 403
        
        print("--- DEBUG ANNOUNCEMENTS PROXY ---")
        print("CONTENT TYPE:", request.content_type)
        print("FORM DATA:", request.form.to_dict())
        print("FILES:", request.files)
        
        if request.content_type and 'multipart' in request.content_type:
            form_data = request.form.to_dict()
            files = {}
            if 'attachment' in request.files:
                attachment_file = request.files['attachment']
                if attachment_file.filename:
                    files['attachment'] = (attachment_file.filename, attachment_file.read(), attachment_file.content_type)
            
            if files:
                print("SENDING MULTIPART TO BACKEND. FILES:", list(files.keys()))
                res = requests.post(f"{BASE_URL}/announcements/", data=form_data, files=files, headers=get_headers(exclude_content_type=True))
            else:
                print("SENDING JSON FALLBACK TO BACKEND:", form_data)
                res = requests.post(f"{BASE_URL}/announcements/", json=form_data, headers=get_headers())
        else:
            print("SENDING JSON TO BACKEND:", request.get_json())
            res = requests.post(f"{BASE_URL}/announcements/", json=request.get_json(), headers=get_headers())
            
        print("BACKEND RESPONSE STATUS:", res.status_code)
        print("BACKEND RESPONSE BODY:", res.content)
        print("---------------------------------")
    else:
        res = requests.get(f"{BASE_URL}/announcements/", params=request.args.to_dict(), headers=get_headers())
    
    try:
        return jsonify(res.json()), res.status_code
    except:
        return Response(res.content, status=res.status_code, headers=dict(res.headers))

@admin_bp.route('/api/announcements/dashboard', methods=['GET'])
@role_required(['admin', 'employee', 'hr', 'manager'])
def api_announcements_dashboard():
    res = requests.get(f"{BASE_URL}/announcements/dashboard", headers=get_headers())
    try:
        return jsonify(res.json()), res.status_code
    except:
        return Response(res.content, status=res.status_code, headers=dict(res.headers))

@admin_bp.route('/api/announcements/<int:announcement_id>', methods=['GET', 'PUT', 'DELETE'])
@role_required(['admin', 'employee', 'hr', 'manager'])
def api_announcement_detail(announcement_id):
    if request.method == 'PUT':
        if str(session.get('role', '')).lower().strip() != 'hr':
            return jsonify({"success": False, "error": "Access denied"}), 403
        
        print("--- DEBUG ANNOUNCEMENT DETAIL PUT PROXY ---")
        print("CONTENT TYPE:", request.content_type)
        print("FORM DATA:", request.form.to_dict())
        print("FILES:", request.files)
        
        if request.content_type and 'multipart' in request.content_type:
            form_data = request.form.to_dict()
            files = {}
            if 'attachment' in request.files:
                attachment_file = request.files['attachment']
                if attachment_file.filename:
                    files['attachment'] = (attachment_file.filename, attachment_file.read(), attachment_file.content_type)
            
            if files:
                print("SENDING MULTIPART PUT TO BACKEND. FILES:", list(files.keys()))
                res = requests.put(f"{BASE_URL}/announcements/{announcement_id}", data=form_data, files=files, headers=get_headers(exclude_content_type=True))
            else:
                print("SENDING JSON PUT FALLBACK TO BACKEND:", form_data)
                res = requests.put(f"{BASE_URL}/announcements/{announcement_id}", json=form_data, headers=get_headers())
        else:
            print("SENDING JSON PUT TO BACKEND:", request.get_json())
            res = requests.put(f"{BASE_URL}/announcements/{announcement_id}", json=request.get_json(), headers=get_headers())
            
        print("BACKEND PUT RESPONSE STATUS:", res.status_code)
        print("BACKEND PUT RESPONSE BODY:", res.content)
        print("--------------------------------------------")
    elif request.method == 'DELETE':
        if str(session.get('role', '')).lower().strip() != 'hr':
            return jsonify({"success": False, "error": "Access denied"}), 403
        res = requests.delete(f"{BASE_URL}/announcements/{announcement_id}", headers=get_headers())
    else:
        res = requests.get(f"{BASE_URL}/announcements/{announcement_id}", headers=get_headers())
        
    try:
        return jsonify(res.json()), res.status_code
    except:
        return Response(res.content, status=res.status_code, headers=dict(res.headers))

@admin_bp.route('/api/announcements/<int:announcement_id>/attachment', methods=['GET'])
@role_required(['admin', 'employee', 'hr', 'manager'])
def api_announcement_attachment(announcement_id):
    res = requests.get(f"{BASE_URL}/announcements/{announcement_id}/attachment", headers=get_headers(), stream=True)
    return Response(res.content, status=res.status_code, headers=dict(res.headers))


# ── RENTAL MANAGEMENT ROUTES ─────────────────────────────────────────────────

@admin_bp.route('/api/rentals/dashboard-stats', methods=['GET'])
@role_required(['admin', 'hr'])
def api_rental_dashboard_stats():
    try:
        res = requests.get(f"{BASE_URL}/rentals/dashboard-stats", params=request.args.to_dict(), headers=get_headers(), timeout=10)
        return jsonify(res.json()), res.status_code
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@admin_bp.route('/api/rentals/matrix', methods=['GET'])
@role_required(['admin', 'hr'])
def api_rental_matrix():
    try:
        res = requests.get(f"{BASE_URL}/rentals/matrix", params=request.args.to_dict(), headers=get_headers(), timeout=15)
        return jsonify(res.json()), res.status_code
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@admin_bp.route('/api/rentals/vendor-summary', methods=['GET'])
@role_required(['admin', 'hr'])
def api_rental_vendor_summary():
    try:
        res = requests.get(f"{BASE_URL}/rentals/vendor-summary", params=request.args.to_dict(), headers=get_headers(), timeout=10)
        return jsonify(res.json()), res.status_code
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@admin_bp.route('/api/rentals/month-summary', methods=['GET'])
@role_required(['admin', 'hr'])
def api_rental_month_summary():
    try:
        res = requests.get(f"{BASE_URL}/rentals/month-summary", params=request.args.to_dict(), headers=get_headers(), timeout=10)
        return jsonify(res.json()), res.status_code
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@admin_bp.route('/api/rentals/export', methods=['GET'])
@role_required(['admin', 'hr'])
def api_rental_export():
    try:
        res = requests.get(f"{BASE_URL}/rentals/export", params=request.args.to_dict(), headers=get_headers(), timeout=60, stream=True)
        if res.status_code == 200:
            return Response(
                res.content,
                status=200,
                headers={
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "Content-Disposition": res.headers.get("Content-Disposition", "attachment; filename=Rental_Report.xlsx")
                }
            )
        return jsonify(res.json()), res.status_code
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
